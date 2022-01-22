#webscraping

import requests, bs4 
from openpyxl import Workbook

def getSoup(link):
    try:
        dataFromSite = requests.get(link)
        soup = bs4.BeautifulSoup(dataFromSite.text, 'lxml')
    except Exception as exc:
        print('There was a problem: %s' % (exc))
    
    return soup

soup = getSoup('https://www.vidal.fr/medicaments/gammes/liste-a.html')

# FETCHING ALL 26 LINKS IN THE ALPHABET RANGE
alphabet_datas = soup.select('.alpha > li > a')
alphabet_links = ['https://vidal.fr/' + link.get('href') for link in alphabet_datas]

# GETTING ALL LINKS IN EACH ALPHABET LETTER
all_links = []
for link in alphabet_links[1:]:    # [1:] to exclude list-0-9 LINK
    soup = getSoup(link)
    data = soup.select('.list > ul > li > a')
    links = ['https://vidal.fr/' + link.get('href') for link in data]

    all_links.extend(links)


# SCRAPPING ALL DATA AFTER GETTING ALL LINKS
all_datas = []

for link in all_links[:2]:
    soup = getSoup(link)
    data = soup.select('.products > .consume-info > ul > li > a')
    links = ['https://vidal.fr/' + link.get('href') for link in data]

    for link in links:
        soup = getSoup(link)
        data1 = soup.select('.packages > .package > .description > .name')  #all 'input' elements that have the attribute 'name' in its attributes
        data2 = soup.select('.packages > .package > .description > .details .cip13')

        alltopics = (data1[0].getText(), data2[0].getText())   #getText() to get text inside html tags (innerHtml)

        all_datas.append(alltopics)

# for i in all_datas:
#     print(i)
    # print('\n')

# CREATING AN EXCEL FILE AND SAVE ALL DATA
wb = Workbook()
sheet = wb.active
# sheet = wb.create_sheet("Mysheet", 0)
# print(wb.sheetnames)

sheet['A1'] = 'Name'
sheet['B1'] = 'cip13'

for i in all_datas:
    sheet.append(i)


wb.save('vidal_data.xlsx')